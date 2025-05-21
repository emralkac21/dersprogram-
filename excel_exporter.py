#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel dışa aktarma modülü
"""

import os
import logging
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    """
    Excel dışa aktarma sınıfı
    """
    
    def __init__(self, db, config):
        """
        Excel dışa aktarma sınıfını başlatır
        
        Args:
            db (Database): Veritabanı bağlantısı
            config (Config): Yapılandırma
        """
        self.db = db
        self.config = config
        self.logger = logging.getLogger(__name__)
        
        # Zaman ayarları
        self.load_time_settings()
        
        # Gün adları
        self.gun_adlari = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
        
        self.logger.info("Excel dışa aktarma sınıfı başlatıldı")
    
    def load_time_settings(self):
        """
        Zaman ayarlarını yükler
        """
        try:
            # Ders süresi
            self.ders_suresi = int(self.db.ayar_getir("ders_suresi", "40"))
            
            # Teneffüs süresi
            self.teneffus_suresi = int(self.db.ayar_getir("teneffus_suresi", "10"))
            
            # Günlük ders başlangıç saati
            self.baslangic_saati = self.db.ayar_getir("gunluk_ders_baslangic", "08:30")
            
            # Günlük ders bitiş saati
            self.bitis_saati = self.db.ayar_getir("gunluk_ders_bitis", "16:00")
            
            # Öğle arası başlangıç saati
            self.ogle_baslangic = self.db.ayar_getir("ogle_arasi_baslangic", "12:00")
            
            # Öğle arası bitiş saati
            self.ogle_bitis = self.db.ayar_getir("ogle_arasi_bitis", "13:00")
            
            # Maksimum günlük ders saati
            self.max_gunluk_ders = int(self.db.ayar_getir("max_gunluk_ders", "8"))
            
            self.logger.info("Zaman ayarları başarıyla yüklendi")
        except Exception as e:
            self.logger.error(f"Zaman ayarları yüklenirken hata oluştu: {str(e)}")
            raise
    
    def calculate_time(self, period_index, teneffus=False):
        """
        Ders saatini hesaplar
        
        Args:
            period_index (int): Ders indeksi
            teneffus (bool): Teneffüs dahil mi?
            
        Returns:
            str: Saat formatında zaman
        """
        try:
            # Başlangıç saatini parse et
            hour, minute = map(int, self.baslangic_saati.split(":"))
            
            # Toplam dakika
            total_minutes = hour * 60 + minute
            
            # Ders indeksine göre dakika ekle
            if period_index > 0:
                for i in range(period_index):
                    # Ders süresi ekle
                    total_minutes += self.ders_suresi
                    
                    # Teneffüs süresi ekle (son ders için teneffüs eklenmez)
                    if i < period_index - 1 or teneffus:
                        total_minutes += self.teneffus_suresi
            
            # Öğle arası kontrolü
            ogle_baslangic_hour, ogle_baslangic_minute = map(int, self.ogle_baslangic.split(":"))
            ogle_bitis_hour, ogle_bitis_minute = map(int, self.ogle_bitis.split(":"))
            
            ogle_baslangic_minutes = ogle_baslangic_hour * 60 + ogle_baslangic_minute
            ogle_bitis_minutes = ogle_bitis_hour * 60 + ogle_bitis_minute
            
            # Eğer hesaplanan zaman öğle arası içindeyse, öğle arası sonrasına kaydır
            if ogle_baslangic_minutes <= total_minutes < ogle_bitis_minutes:
                total_minutes = ogle_bitis_minutes
            
            # Dakikayı saat ve dakika olarak dönüştür
            hour = total_minutes // 60
            minute = total_minutes % 60
            
            # Saat formatında döndür
            return f"{hour:02d}:{minute:02d}"
        except Exception as e:
            self.logger.error(f"Saat hesaplanırken hata oluştu: {str(e)}")
            return "00:00"
    
    def get_ders_color(self, ders_id):
        """
        Ders ID'sine göre renk döndürür
        
        Args:
            ders_id (int): Ders ID'si
            
        Returns:
            str: Renk kodu (Excel formatında)
        """
        # Ders ID'sine göre sabit bir renk üret
        colors = [
            "FFD6D6", "D6FFD6", "D6D6FF", "FFFFD6", "FFD6FF", "D6FFFF",
            "FFE0D6", "D6FFE0", "E0D6FF", "FFFFE0", "FFE0FF", "E0FFFF"
        ]
        
        return colors[ders_id % len(colors)]
    
    def export_sinif_programi(self, sinif_id, output_path):
        """
        Sınıf programını Excel olarak dışa aktarır
        
        Args:
            sinif_id (int): Sınıf ID'si
            output_path (str): Çıktı dosya yolu
            
        Returns:
            bool: Başarılı mı?
        """
        try:
            # Sınıf bilgilerini al
            self.db.execute("SELECT * FROM siniflar WHERE id = ?", (sinif_id,))
            sinif = self.db.fetchone()
            
            if not sinif:
                self.logger.error(f"Sınıf bulunamadı: {sinif_id}")
                return False
            
            # Program verilerini al
            self.db.execute("""
                SELECT p.*, s.ad as sinif_adi, s.sube as sinif_sube, 
                       o.ad_soyad as ogretmen_adi, d.ad as ders_adi, 
                       dr.ad as derslik_adi, d.id as ders_id
                FROM program p
                JOIN siniflar s ON p.sinif_id = s.id
                JOIN ogretmenler o ON p.ogretmen_id = o.id
                JOIN dersler d ON p.ders_id = d.id
                JOIN derslikler dr ON p.derslik_id = dr.id
                WHERE p.sinif_id = ?
                ORDER BY p.gun, p.saat
            """, (sinif_id,))
            
            program_verileri = self.db.fetchall()
            
            # Program verilerini gün ve saate göre düzenle
            program_tablosu = {}
            for gun in range(len(self.gun_adlari)):
                program_tablosu[gun] = {}
                for saat in range(self.max_gunluk_ders):
                    program_tablosu[gun][saat] = None
            
            for ders in program_verileri:
                program_tablosu[ders["gun"]][ders["saat"]] = ders
            
            # Excel çalışma kitabı oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{sinif['ad']} {sinif['sube']}"
            
            # Başlık ve üst bilgi
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Sınıf Programı - {sinif['ad']} {sinif['sube']}"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A2:G2')
            ws['A2'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Tablo başlıkları
            ws['A4'] = "Saat / Gün"
            ws['A4'].font = Font(bold=True)
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A4'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Gün başlıkları
            for i, gun_adi in enumerate(self.gun_adlari):
                col = get_column_letter(i + 2)
                ws[f'{col}4'] = gun_adi
                ws[f'{col}4'].font = Font(bold=True)
                ws[f'{col}4'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}4'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Ders saatleri ve program
            for saat in range(self.max_gunluk_ders):
                row = saat + 5
                saat_baslangic = self.calculate_time(saat)
                saat_bitis = self.calculate_time(saat + 1, teneffus=True)
                
                # Saat hücresi
                ws[f'A{row}'] = f"{saat+1}. Ders\n{saat_baslangic}-{saat_bitis}"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                
                # Her gün için
                for gun in range(len(self.gun_adlari)):
                    col = get_column_letter(gun + 2)
                    ders = program_tablosu[gun][saat]
                    
                    if ders:
                        # Ders bilgisi
                        ws[f'{col}{row}'] = f"{ders['ders_adi']}\n{ders['ogretmen_adi']}\n{ders['derslik_adi']}"
                        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Ders rengi
                        bg_color = self.get_ders_color(ders["ders_id"])
                        ws[f'{col}{row}'].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            # Tablo kenarlıkları
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in range(4, self.max_gunluk_ders + 5):
                for col in range(1, len(self.gun_adlari) + 2):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            for i in range(len(self.gun_adlari)):
                col = get_column_letter(i + 2)
                ws.column_dimensions[col].width = 20
            
            # Satır yükseklikleri
            for i in range(5, self.max_gunluk_ders + 5):
                ws.row_dimensions[i].height = 60
            
            # Excel dosyasını kaydet
            wb.save(output_path)
            
            self.logger.info(f"Sınıf programı Excel olarak dışa aktarıldı: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"Sınıf programı dışa aktarılırken hata oluştu: {str(e)}")
            raise
    
    def export_ogretmen_programi(self, ogretmen_id, output_path):
        """
        Öğretmen programını Excel olarak dışa aktarır
        
        Args:
            ogretmen_id (int): Öğretmen ID'si
            output_path (str): Çıktı dosya yolu
            
        Returns:
            bool: Başarılı mı?
        """
        try:
            # Öğretmen bilgilerini al
            self.db.execute("SELECT * FROM ogretmenler WHERE id = ?", (ogretmen_id,))
            ogretmen = self.db.fetchone()
            
            if not ogretmen:
                self.logger.error(f"Öğretmen bulunamadı: {ogretmen_id}")
                return False
            
            # Program verilerini al
            self.db.execute("""
                SELECT p.*, s.ad as sinif_adi, s.sube as sinif_sube, 
                       o.ad_soyad as ogretmen_adi, d.ad as ders_adi, 
                       dr.ad as derslik_adi, d.id as ders_id
                FROM program p
                JOIN siniflar s ON p.sinif_id = s.id
                JOIN ogretmenler o ON p.ogretmen_id = o.id
                JOIN dersler d ON p.ders_id = d.id
                JOIN derslikler dr ON p.derslik_id = dr.id
                WHERE p.ogretmen_id = ?
                ORDER BY p.gun, p.saat
            """, (ogretmen_id,))
            
            program_verileri = self.db.fetchall()
            
            # Program verilerini gün ve saate göre düzenle
            program_tablosu = {}
            for gun in range(len(self.gun_adlari)):
                program_tablosu[gun] = {}
                for saat in range(self.max_gunluk_ders):
                    program_tablosu[gun][saat] = None
            
            for ders in program_verileri:
                program_tablosu[ders["gun"]][ders["saat"]] = ders
            
            # Excel çalışma kitabı oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = ogretmen['ad_soyad']
            
            # Başlık ve üst bilgi
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Öğretmen Programı - {ogretmen['ad_soyad']}"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A2:G2')
            ws['A2'] = f"Branş: {ogretmen['brans']}"
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A3:G3')
            ws['A3'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Tablo başlıkları
            ws['A5'] = "Saat / Gün"
            ws['A5'].font = Font(bold=True)
            ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A5'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Gün başlıkları
            for i, gun_adi in enumerate(self.gun_adlari):
                col = get_column_letter(i + 2)
                ws[f'{col}5'] = gun_adi
                ws[f'{col}5'].font = Font(bold=True)
                ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}5'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Ders saatleri ve program
            for saat in range(self.max_gunluk_ders):
                row = saat + 6
                saat_baslangic = self.calculate_time(saat)
                saat_bitis = self.calculate_time(saat + 1, teneffus=True)
                
                # Saat hücresi
                ws[f'A{row}'] = f"{saat+1}. Ders\n{saat_baslangic}-{saat_bitis}"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                
                # Her gün için
                for gun in range(len(self.gun_adlari)):
                    col = get_column_letter(gun + 2)
                    ders = program_tablosu[gun][saat]
                    
                    if ders:
                        # Ders bilgisi
                        ws[f'{col}{row}'] = f"{ders['ders_adi']}\n{ders['sinif_adi']} {ders['sinif_sube']}\n{ders['derslik_adi']}"
                        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Ders rengi
                        bg_color = self.get_ders_color(ders["ders_id"])
                        ws[f'{col}{row}'].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            # Tablo kenarlıkları
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in range(5, self.max_gunluk_ders + 6):
                for col in range(1, len(self.gun_adlari) + 2):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            for i in range(len(self.gun_adlari)):
                col = get_column_letter(i + 2)
                ws.column_dimensions[col].width = 20
            
            # Satır yükseklikleri
            for i in range(6, self.max_gunluk_ders + 6):
                ws.row_dimensions[i].height = 60
            
            # Excel dosyasını kaydet
            wb.save(output_path)
            
            self.logger.info(f"Öğretmen programı Excel olarak dışa aktarıldı: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"Öğretmen programı dışa aktarılırken hata oluştu: {str(e)}")
            raise
    
    def export_derslik_programi(self, derslik_id, output_path):
        """
        Derslik programını Excel olarak dışa aktarır
        
        Args:
            derslik_id (int): Derslik ID'si
            output_path (str): Çıktı dosya yolu
            
        Returns:
            bool: Başarılı mı?
        """
        try:
            # Derslik bilgilerini al
            self.db.execute("SELECT * FROM derslikler WHERE id = ?", (derslik_id,))
            derslik = self.db.fetchone()
            
            if not derslik:
                self.logger.error(f"Derslik bulunamadı: {derslik_id}")
                return False
            
            # Program verilerini al
            self.db.execute("""
                SELECT p.*, s.ad as sinif_adi, s.sube as sinif_sube, 
                       o.ad_soyad as ogretmen_adi, d.ad as ders_adi, 
                       dr.ad as derslik_adi, d.id as ders_id
                FROM program p
                JOIN siniflar s ON p.sinif_id = s.id
                JOIN ogretmenler o ON p.ogretmen_id = o.id
                JOIN dersler d ON p.ders_id = d.id
                JOIN derslikler dr ON p.derslik_id = dr.id
                WHERE p.derslik_id = ?
                ORDER BY p.gun, p.saat
            """, (derslik_id,))
            
            program_verileri = self.db.fetchall()
            
            # Program verilerini gün ve saate göre düzenle
            program_tablosu = {}
            for gun in range(len(self.gun_adlari)):
                program_tablosu[gun] = {}
                for saat in range(self.max_gunluk_ders):
                    program_tablosu[gun][saat] = None
            
            for ders in program_verileri:
                program_tablosu[ders["gun"]][ders["saat"]] = ders
            
            # Excel çalışma kitabı oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = derslik['ad']
            
            # Başlık ve üst bilgi
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Derslik Programı - {derslik['ad']}"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A2:G2')
            ws['A2'] = f"Tür: {derslik['tur']}"
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A3:G3')
            ws['A3'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Tablo başlıkları
            ws['A5'] = "Saat / Gün"
            ws['A5'].font = Font(bold=True)
            ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
            ws['A5'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Gün başlıkları
            for i, gun_adi in enumerate(self.gun_adlari):
                col = get_column_letter(i + 2)
                ws[f'{col}5'] = gun_adi
                ws[f'{col}5'].font = Font(bold=True)
                ws[f'{col}5'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}5'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Ders saatleri ve program
            for saat in range(self.max_gunluk_ders):
                row = saat + 6
                saat_baslangic = self.calculate_time(saat)
                saat_bitis = self.calculate_time(saat + 1, teneffus=True)
                
                # Saat hücresi
                ws[f'A{row}'] = f"{saat+1}. Ders\n{saat_baslangic}-{saat_bitis}"
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                
                # Her gün için
                for gun in range(len(self.gun_adlari)):
                    col = get_column_letter(gun + 2)
                    ders = program_tablosu[gun][saat]
                    
                    if ders:
                        # Ders bilgisi
                        ws[f'{col}{row}'] = f"{ders['ders_adi']}\n{ders['sinif_adi']} {ders['sinif_sube']}\n{ders['ogretmen_adi']}"
                        ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Ders rengi
                        bg_color = self.get_ders_color(ders["ders_id"])
                        ws[f'{col}{row}'].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
            # Tablo kenarlıkları
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in range(5, self.max_gunluk_ders + 6):
                for col in range(1, len(self.gun_adlari) + 2):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Sütun genişlikleri
            ws.column_dimensions['A'].width = 15
            for i in range(len(self.gun_adlari)):
                col = get_column_letter(i + 2)
                ws.column_dimensions[col].width = 20
            
            # Satır yükseklikleri
            for i in range(6, self.max_gunluk_ders + 6):
                ws.row_dimensions[i].height = 60
            
            # Excel dosyasını kaydet
            wb.save(output_path)
            
            self.logger.info(f"Derslik programı Excel olarak dışa aktarıldı: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"Derslik programı dışa aktarılırken hata oluştu: {str(e)}")
            raise
    
    def export_tum_sinif_programlari(self, output_dir):
        """
        Tüm sınıf programlarını Excel olarak dışa aktarır
        
        Args:
            output_dir (str): Çıktı dizini
            
        Returns:
            bool: Başarılı mı?
        """
        try:
            # Sınıfları al
            self.db.execute("SELECT * FROM siniflar ORDER BY ad, sube")
            siniflar = self.db.fetchall()
            
            if not siniflar:
                self.logger.error("Hiç sınıf bulunamadı")
                return False
            
            # Her sınıf için Excel oluştur
            for sinif in siniflar:
                output_path = os.path.join(output_dir, f"sinif_programi_{sinif['ad']}_{sinif['sube']}.xlsx")
                self.export_sinif_programi(sinif["id"], output_path)
            
            self.logger.info(f"Tüm sınıf programları Excel olarak dışa aktarıldı: {output_dir}")
            return True
        except Exception as e:
            self.logger.error(f"Tüm sınıf programları dışa aktarılırken hata oluştu: {str(e)}")
            raise
    
    def export_tum_ogretmen_programlari(self, output_dir):
        """
        Tüm öğretmen programlarını Excel olarak dışa aktarır
        
        Args:
            output_dir (str): Çıktı dizini
            
        Returns:
            bool: Başarılı mı?
        """
        try:
            # Öğretmenleri al
            self.db.execute("SELECT * FROM ogretmenler ORDER BY ad_soyad")
            ogretmenler = self.db.fetchall()
            
            if not ogretmenler:
                self.logger.error("Hiç öğretmen bulunamadı")
                return False
            
            # Her öğretmen için Excel oluştur
            for ogretmen in ogretmenler:
                output_path = os.path.join(output_dir, f"ogretmen_programi_{ogretmen['ad_soyad'].replace(' ', '_')}.xlsx")
                self.export_ogretmen_programi(ogretmen["id"], output_path)
            
            self.logger.info(f"Tüm öğretmen programları Excel olarak dışa aktarıldı: {output_dir}")
            return True
        except Exception as e:
            self.logger.error(f"Tüm öğretmen programları dışa aktarılırken hata oluştu: {str(e)}")
            raise
    
    def export_tum_derslik_programlari(self, output_dir):
        """
        Tüm derslik programlarını Excel olarak dışa aktarır
        
        Args:
            output_dir (str): Çıktı dizini
            
        Returns:
            bool: Başarılı mı?
        """
        try:
            # Derslikleri al
            self.db.execute("SELECT * FROM derslikler ORDER BY ad")
            derslikler = self.db.fetchall()
            
            if not derslikler:
                self.logger.error("Hiç derslik bulunamadı")
                return False
            
            # Her derslik için Excel oluştur
            for derslik in derslikler:
                output_path = os.path.join(output_dir, f"derslik_programi_{derslik['ad'].replace(' ', '_')}.xlsx")
                self.export_derslik_programi(derslik["id"], output_path)
            
            self.logger.info(f"Tüm derslik programları Excel olarak dışa aktarıldı: {output_dir}")
            return True
        except Exception as e:
            self.logger.error(f"Tüm derslik programları dışa aktarılırken hata oluştu: {str(e)}")
            raise
